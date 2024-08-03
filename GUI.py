import subprocess
import time
import tkinter
import customtkinter
import functools
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles.numbers import FORMAT_NUMBER_00

#System settings
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

#Load BaseBuildingSheet and BuildingSheet + initialize some arrays

buildings = []

baseBSWB = load_workbook(filename='BaseBuildingSheet.xlsx')
buildingNames = baseBSWB.sheetnames

BSWB = load_workbook(filename='BuildingSheet.xlsx')
BS = BSWB['OptimizationTable']

#App frame

app = customtkinter.CTk()
app.geometry("1280x1080")
app.resizable(False, False)
app.title("Building efficiency optimizer")

#Functions

def AddBuilding():
    
    place = addBuildingButton.grid_info()['row']

    addBuildingButton.grid(row=place+1, column=0)
    newRow = []
    #Adding button for deleting a building from calculations
    newRow.append(customtkinter.CTkButton(scrollFrame, text='Delete' , width=10, bg_color='red', command=functools.partial(DeleteBuildingRow, place)))
    newRow[0].grid(row=place, column=0, padx=(2,10), pady=10)
    #Adding combobox that adds a buildings to calculations
    newRow.append(customtkinter.CTkComboBox(scrollFrame, values=buildingNames, width=100, state='readonly', command=functools.partial(AddBuildingPMs, place)))
    newRow[1].grid(row=place, column=1, padx=10, pady=10)
    buildings.append(newRow)
    scrollFrame.after(0, scrollFrame._parent_canvas.yview_moveto, 1.0)
    

#Function for removing a building from calculations
def DeleteBuildingRow(rowNumber):

    while len(buildings[rowNumber-1]) > 0:
        buildings[rowNumber-1][0].destroy()
        buildings[rowNumber-1].pop(0)

#Adds selectable PMs for a given type of building
def AddBuildingPMs(rowNumber, buildingName):
    baseBS = baseBSWB[buildingName]

    pmOptions = ParsePMOptions(baseBS.data_validations.dataValidation)
    i = 1
    #Empty previous PMs if row already exists
    if len(buildings[rowNumber-1]) != 2:
        while len(buildings[rowNumber-1]) > 2:
            buildings[rowNumber-1][i+1].destroy()
            buildings[rowNumber-1].pop(i+1)
    #Add choices for each PM
    for pm in pmOptions:
        buildings[rowNumber-1].append(customtkinter.CTkComboBox(scrollFrame, width=75, state='readonly', values=pm))
        i += 1
        if(len(pm)>0):
            buildings[rowNumber-1][i].set(pm[0])
        buildings[rowNumber-1][i].grid(row=rowNumber, column=i, padx=5, pady=10)
        
        
    
    #Entry for throughput bonus
    buildings[rowNumber-1].append(customtkinter.CTkEntry(scrollFrame, width=50))
    i += 1
    buildings[rowNumber-1][i].grid(row=rowNumber, column=i, padx=5, pady=10)
    
    #Entry for construction bonus
    buildings[rowNumber-1].append(customtkinter.CTkEntry(scrollFrame, width=50))
    i += 1
    buildings[rowNumber-1][i].grid(row=rowNumber, column=i, padx=5, pady=10)

#Function for parsing the information of a DataValidation object into something we can build UI elements from
def ParsePMOptions(validatedData):
    options = [[]] * 4
    for entry in validatedData:
        slot = entry.sqref
        if slot == 'B2':
            options[0] = Formula1Parser(entry.formula1)
        if slot == 'C2':
            options[1] = Formula1Parser(entry.formula1)
        if slot == 'D2':
            options[2] = Formula1Parser(entry.formula1)
        if slot == 'E2':
            options[3] = Formula1Parser(entry.formula1)
    return options

#Function for parsing a formula1 string of a DataValidation object into a list of elements
def Formula1Parser(formula):
    return formula.replace('\"', '').split(',')    

#Function for slecting correct row form BaseBuildingSheet
def FindRow(val, sheet):
    for i in range(3,sheet.max_row):
        if sheet.cell(row=i, column=2).value==val:
            return i
    return 2

#Setup BuildingSheet for calculations
def BEOFunctionality():

    #Emptying BuildingSheet
    BS.delete_rows(2,BS.max_row)
    BSWB.save('BuildingSheet.xlsx')

    #Adding selected buildings and PMs to the BuildingSheet
    r=2
    for row in buildings:
        if len(row) > 0:
            #Select correct building
            entry = baseBSWB[row[1].get()]

            #Add base values for Construction, Labor and Inputs/Outputs
            for j in range(9,99):
                BS.cell(row=r, column=j+1).value = entry.cell(row=2, column=j).value
            BSWB.save('BuildingSheet.xlsx')            
            i=1
            #Add impact of PMs to Labor and Inputs/Outputs
            for i in range(2,len(row)-2):
                val = row[i].get()
                if val != '':
                    pm = FindRow(val, entry)
                    for j in range(10,99):
                        BS.cell(row=r, column=j+1).value = BS.cell(row=r, column=j+1).value + entry.cell(row=pm, column=j).value
            BSWB.save('BuildingSheet.xlsx')
            #Including building in calculations
            BS.cell(row=r, column=7).value=1
            #Add throughput bonus and construction bonus
            for i in range(i, len(row)-1):
                try:
                    val=float(row[i].get())
                    BS.cell(row=r, column=i+3).value=val/100
                except:
                    BS.cell(row=r, column=i+3).value=0
                BS.cell(row=r, column=i+3).number_format = FORMAT_PERCENTAGE_00
            BSWB.save('BuildingSheet.xlsx')
            r += 1
            
    #Calculate the prices!
    CalculateOptimalPrices()

#Launching OptimizeBuildings.py and printing status messages to statusLabel
def CalculateOptimalPrices():

    pricingProcess = subprocess.Popen(['python', 'OptimizeBuildings.py'], stdout=subprocess.PIPE, bufsize=1, universal_newlines=True)
    for line in pricingProcess.stdout:
        app.update()
        currentText = statusText.get()
        if currentText.count('\n')>8:
            currentText = currentText.split('\n',currentText.count('\n')-8)[currentText.count('\n')-8]
        statusText.set(currentText+'\n'+line)
        app.update()

    
#Showing prices in the OptimizedPrices worksheet
def ShowOptimalPrices(type):
    try:
        priceWB = load_workbook(filename='OptimizedPrices.xlsx')
        priceSheet = priceWB['Optimized for construction']
        if(type == "Lab"):
            priceSheet = priceWB['Optimized for labor']
        
        r=1
        #Adding all gods that were affected by the calculations
        for i in range(2,priceSheet.max_row):
            if priceSheet.cell(row=i, column=7).value is None:
                #Good name
                good=priceSheet.cell(row=i, column=2).value
                label1 = customtkinter.CTkLabel(priceFrame, width=50, height=50, text=good, bg_color="white")
                label1.grid(row=r, column=0, padx=10, pady=(10, 20))
                #Absolute optimal price
                absPrice=round(priceSheet.cell(row=i, column=4).value,1)
                label2 = customtkinter.CTkLabel(priceFrame, width=50, height=50, text=absPrice, bg_color="white")
                label2.grid(row=r, column=1, padx=10, pady=(10, 20))
                #Absolute price relative to base price
                relPrice=priceSheet.cell(row=i, column=5).value
                label3 = customtkinter.CTkLabel(priceFrame, width=50, height=50, text=relPrice, bg_color="white")
                label3.grid(row=r, column=2, padx=10, pady=(10, 20))
            r+=1
            
    except:
        statusText.set("Optimized prices do not yet exits. Please perform a calculation to show prices!")

    

#UI elements

#Scrollable frame for buildings
scrollFrame = customtkinter.CTkScrollableFrame(app, orientation="vertical", width=800, height=1080)
scrollFrame.pack(padx=(0,25), side=customtkinter.LEFT, anchor=customtkinter.NW)
#Fix for scrolling not working on Linux
scrollFrame.bind("<Button-4>", lambda e: scrollFrame._parent_canvas.yview("scroll", -1, "units"))
scrollFrame.bind("<Button-5>", lambda e: scrollFrame._parent_canvas.yview("scroll", 1, "units"))
scrollFrame._scrollbar.configure(height=0)

statusText = customtkinter.StringVar()
statusText.set("Waiting for launch")
statusFont = customtkinter.CTkFont(family="HELVETICA", size=14)
statusLabel = customtkinter.CTkLabel(app, textvariable=statusText, width=300, height=150, font=statusFont)
statusLabel.pack(padx=10, pady = 50)

calculateButton = customtkinter.CTkButton(app, text="Calculate prices", width=150, height=50, command=BEOFunctionality)
calculateButton.pack(padx=50, pady=10)


#Scrollable frame for prices
priceFrame = customtkinter.CTkScrollableFrame(app, orientation="vertical", width=400, height=800)
priceFrame.pack(padx=50, pady=10)
priceFrame.bind("<Button-4>", lambda e: scrollFrame._parent_canvas.yview("scroll", -1, "units"))
priceFrame.bind("<Button-5>", lambda e: scrollFrame._parent_canvas.yview("scroll", 1, "units"))
priceFrame._scrollbar.configure(height=0)

goodLabel = customtkinter.CTkLabel(priceFrame, width=75, height=50, text="Good: ")
goodLabel.grid(row=0, column=0, padx=10)

priceLabel = customtkinter.CTkLabel(priceFrame, width=50, height=50, text="Price: ")
priceLabel.grid(row=0, column=1, padx=10)

precentLabel = customtkinter.CTkLabel(priceFrame, width=50, height=50, text="%: ")
precentLabel.grid(row=0, column=2, padx=10)

conButton = customtkinter.CTkButton(priceFrame, text='Con' , width=10, command=functools.partial(ShowOptimalPrices, "Con"))
conButton.grid(row=0, column=3)
    
labButton = customtkinter.CTkButton(priceFrame, text='Lab' , width=10, command=functools.partial(ShowOptimalPrices, "Lab"))
labButton.grid(row=0, column=4)

#Button for adding buildings to calculations
addBuildingButton = customtkinter.CTkButton(scrollFrame, width=10, height=10, text="Add Building", command=AddBuilding)
addBuildingButton.grid(row=1, column=0, pady = (20, 650))

#Labels for rows of buildings

typeLabel = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="Building Type")
typeLabel.grid(row=0, column=1, padx=10)

pmLabel1 = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="PM #1")
pmLabel1.grid(row=0, column=2, padx=10)

pmLabel2 = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="PM #2")
pmLabel2.grid(row=0, column=3, padx=10)

pmLabel3 = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="PM #3")
pmLabel3.grid(row=0, column=4, padx=10)

pmLabel4 = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="PM #4")
pmLabel4.grid(row=0, column=5, padx=10)


throughputLabel = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="TPutBonus %")
throughputLabel.grid(row=0, column=6, padx=5)

constructionLabel = customtkinter.CTkLabel(scrollFrame, width=50, height=50, text="ConBonus %")
constructionLabel.grid(row=0, column=7)




# Run app

app.mainloop()